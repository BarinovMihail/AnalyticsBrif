from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import CardCharacteristic, MTRCard
from app.services.identifier_utils import extract_identifier, normalize_text


logger = logging.getLogger(__name__)


GUID_RE = re.compile(r"\b[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\b")
INN_RE = re.compile(r"\b\d{10,12}\b")
DATE_RE = re.compile(r"\b\d{4}\.\d{2}\.\d{2}\b")
RANGE_RE = re.compile(r"([+-]?\d+(?:[.,]\d+)?)\s*(?:\.\.\.|/)\s*([+-]?\d+(?:[.,]\d+)?)")
META_HEADERS = {
    "guid",
    "наименование",
    "инн изготовителя",
    "артикул",
    "цена exw без ндс",
    "дата начала действия цены",
    "плановая дата окончания действия цены",
    "описание",
    "пункт спецперечня",
}


def _normalize_text(value: object) -> str | None:
    return normalize_text(value)


def _clean_char_name(value: str | None) -> str | None:
    if not value:
        return value
    cut = re.split(r"\s+тип\s+данных", value, maxsplit=1, flags=re.IGNORECASE)
    return cut[0].strip() if cut else value.strip()


def _to_float(value: str) -> float | None:
    try:
        return float(value.replace(" ", "").replace(",", "."))
    except (AttributeError, ValueError):
        return None


def _to_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(value.replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y.%m.%d").date()
    except ValueError:
        return None


def _row_values(row) -> list[str]:
    return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]


def _normalize_header_name(value: object) -> str:
    text = _normalize_text(value) or ""
    return " ".join(text.replace("\n", " ").split()).strip().lower()


def _find_header_index(headers: list[object], needle: str) -> int | None:
    normalized_needle = needle.lower()
    for index, header in enumerate(headers):
        normalized_header = _normalize_header_name(header)
        if normalized_needle in normalized_header:
            return index
    return None


def _extract_header(row_values: list[str]) -> dict | None:
    combined = " | ".join(row_values)
    guid_match = GUID_RE.search(combined)
    if not guid_match:
        return None

    guid = guid_match.group(0)
    manufacturer_inn = next((value for value in row_values if INN_RE.fullmatch(value)), None)
    manufacturer_inio = next(
        (
            value
            for value in row_values
            if value
            and not GUID_RE.fullmatch(value)
            and not DATE_RE.fullmatch(value)
            and not INN_RE.fullmatch(value)
            and re.fullmatch(r"(?=.*[A-Za-zА-Яа-я])[A-Za-zА-Яа-я0-9]+", value)
        ),
        None,
    )
    identifier = extract_identifier(manufacturer_inn, inio_value=manufacturer_inio)
    manufacturer_marker = manufacturer_inn or manufacturer_inio
    delivery_date = next((value for value in row_values if DATE_RE.fullmatch(value)), None)
    price = None

    candidate_numbers = []
    for value in row_values:
        normalized = value.replace(" ", "").replace(",", ".")
        if normalized == manufacturer_inn or normalized == delivery_date:
            continue
        try:
            float(normalized)
            candidate_numbers.append(value)
        except ValueError:
            continue
    if candidate_numbers:
        price = candidate_numbers[-1]

    filtered = [value for value in row_values if value != "TITLE"]
    guid_index = next((i for i, value in enumerate(filtered) if guid in value), 0)
    name = filtered[guid_index + 1] if len(filtered) > guid_index + 1 else None
    if not name:
        non_meta = [
            value
            for value in filtered
            if value != guid
            and value != manufacturer_inn
            and value != delivery_date
            and value != price
        ]
        name = non_meta[0] if non_meta else None

    mtr_class = None
    if manufacturer_marker:
        try:
            manufacturer_index = filtered.index(manufacturer_marker)
            if manufacturer_index + 1 < len(filtered):
                mtr_class = filtered[manufacturer_index + 1]
        except ValueError:
            mtr_class = None

    currency_code = None
    for value in filtered:
        if re.fullmatch(r"\d{3}", value):
            currency_code = value
            break

    return {
        "guid": guid,
        "nomenclature_name": name,
        "manufacturer_inn": identifier["value"],
        "manufacturer_identifier_type": identifier["type"],
        "mtr_class": mtr_class,
        "price": _to_decimal(price),
        "currency_code": currency_code,
        "delivery_date": _parse_date(delivery_date),
    }


def _parse_characteristic_value(raw_value: str | None) -> tuple[str, float | None, float | None, float | None]:
    value = raw_value or ""
    numeric_value = _to_float(value)
    range_min = None
    range_max = None

    match = RANGE_RE.search(value)
    if match:
        range_min = _to_float(match.group(1))
        range_max = _to_float(match.group(2))

    return value, numeric_value, range_min, range_max


def _detect_file2_format(sheet) -> str:
    row1 = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    row3 = [cell for cell in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))]

    normalized_row1 = [_normalize_header_name(value) for value in row1]
    normalized_row3 = [_normalize_header_name(value) for value in row3]

    if "класс мтр" in normalized_row1 and "guid" in normalized_row3 and "наименование" in normalized_row3:
        return "tabular"
    return "legacy"


def _import_tabular_file2(db: Session, workbook) -> dict:
    sheet = workbook.active
    headers = [cell for cell in next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))]
    mtr_class_row = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    default_mtr_class = _normalize_text(mtr_class_row[1]) if len(mtr_class_row) > 1 else None

    guid_index = _find_header_index(headers, "guid")
    nomenclature_index = _find_header_index(headers, "наименование")
    inn_index = _find_header_index(headers, "инн изготовителя")
    inio_index = _find_header_index(headers, "инио изготовителя")
    price_index = _find_header_index(headers, "цена exw без ндс")
    start_date_index = _find_header_index(headers, "дата начала действия цены")
    plan_date_index = _find_header_index(headers, "плановая дата окончания действия цены")
    currency_index = _find_header_index(headers, "валюта")

    meta_indexes = {
        index
        for index in [
            guid_index,
            nomenclature_index,
            inn_index,
            inio_index,
            price_index,
            start_date_index,
            plan_date_index,
            currency_index,
        ]
        if index is not None
    }
    meta_indexes.update(
        index
        for index, header in enumerate(headers)
        if _normalize_header_name(header) in META_HEADERS
        or "страна регистрации изготовителя" in _normalize_header_name(header)
    )

    characteristic_indexes = [
        index
        for index, header in enumerate(headers)
        if index not in meta_indexes and _normalize_text(header)
    ]

    existing_guids = set(db.scalars(select(MTRCard.guid)).all())
    imported = 0
    skipped = 0
    errors: list[str] = []
    skipped_rows: list[dict] = []

    def _record_skip(row_num: int, reason: str, preview: dict[str, str] | None = None) -> None:
        skipped_rows.append({"row": row_num, "reason": reason, "preview": preview or {}})

    for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(row)
        try:
            guid = _normalize_text(values[guid_index]) if guid_index is not None and guid_index < len(values) else None
            if not guid:
                if all(_normalize_text(value) is None for value in values):
                    skipped += 1
                    _record_skip(row_index, "Пустая строка")
                    continue
                logger.warning("Строка %s: отсутствует GUID, строка пропущена", row_index)
                skipped += 1
                _record_skip(row_index, "Отсутствует GUID")
                continue

            if guid in existing_guids:
                skipped += 1
                _record_skip(row_index, "Дубликат GUID", {"GUID": guid})
                continue

            manufacturer_name = None
            identifier = extract_identifier(
                values[inn_index] if inn_index is not None and inn_index < len(values) else None,
                name_value=manufacturer_name,
                inio_value=values[inio_index] if inio_index is not None and inio_index < len(values) else None,
            )
            card = MTRCard(
                guid=guid,
                nomenclature_name=_normalize_text(values[nomenclature_index]) if nomenclature_index is not None and nomenclature_index < len(values) else None,
                manufacturer_inn=identifier["value"],
                manufacturer_identifier_type=identifier["type"],
                mtr_class=default_mtr_class,
                price=_to_decimal(_normalize_text(values[price_index])) if price_index is not None and price_index < len(values) else None,
                currency_code=_normalize_text(values[currency_index]) if currency_index is not None and currency_index < len(values) else None,
                delivery_date=_parse_date(_normalize_text(values[plan_date_index])) if plan_date_index is not None and plan_date_index < len(values) else None,
            )

            for index in characteristic_indexes:
                if index >= len(values):
                    continue
                char_name = _clean_char_name(_normalize_text(headers[index]))
                raw_value = _normalize_text(values[index])
                if not char_name or not raw_value:
                    continue
                char_value, numeric_value, range_min, range_max = _parse_characteristic_value(raw_value)
                card.characteristics.append(
                    CardCharacteristic(
                        char_name=char_name,
                        char_value=char_value,
                        char_value_numeric=numeric_value,
                        range_min=range_min,
                        range_max=range_max,
                    )
                )

            db.add(card)
            existing_guids.add(guid)
            imported += 1
        except Exception as exc:
            message = f"Строка {row_index}: {exc}"
            logger.exception(message)
            errors.append(message)

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors, "skipped_rows": skipped_rows}


def import_file2(db: Session, file_path: str | Path) -> dict:
    workbook = load_workbook(filename=file_path, data_only=True)
    if _detect_file2_format(workbook.active) == "tabular":
        return _import_tabular_file2(db, workbook)

    sheet = workbook.active

    existing_guids = set(db.scalars(select(MTRCard.guid)).all())
    imported = 0
    skipped = 0
    errors: list[str] = []
    skipped_rows: list[dict] = []

    def _record_skip(row_num: int, reason: str, preview: dict[str, str] | None = None) -> None:
        skipped_rows.append({"row": row_num, "reason": reason, "preview": preview or {}})

    current_card: MTRCard | None = None

    def flush_card(card: MTRCard | None):
        nonlocal imported
        if card is None:
            return
        db.add(card)
        imported += 1

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = _row_values(row)
        if not values:
            continue

        try:
            is_header = "TITLE" in values or any(GUID_RE.search(value) for value in values)
            if is_header:
                header = _extract_header(values)
                if header is None:
                    raise ValueError("Не удалось определить GUID заголовка блока")

                guid = header["guid"]
                if guid in existing_guids:
                    skipped += 1
                    _record_skip(row_index, "Дубликат GUID", {"GUID": guid})
                    current_card = None
                    continue

                flush_card(current_card)
                current_card = MTRCard(**header)
                existing_guids.add(guid)
                continue

            if current_card is None:
                continue

            if len(values) < 2:
                raise ValueError("Строка характеристики должна содержать char_name и char_value")

            char_name = _clean_char_name(values[0])
            char_value, numeric_value, range_min, range_max = _parse_characteristic_value(values[1])
            current_card.characteristics.append(
                CardCharacteristic(
                    char_name=char_name,
                    char_value=char_value,
                    char_value_numeric=numeric_value,
                    range_min=range_min,
                    range_max=range_max,
                )
            )
        except Exception as exc:
            message = f"Строка {row_index}: {exc}"
            logger.exception(message)
            errors.append(message)

    flush_card(current_card)
    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors, "skipped_rows": skipped_rows}
